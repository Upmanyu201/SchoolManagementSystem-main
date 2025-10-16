# backup/services/export_service.py
# Minimal export  for School Management System
import csv
import io
import logging
import time
import os
from datetime import datetime
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_protect
from django.core.cache import cache
from django.utils._os import safe_join
from django.core.exceptions import PermissionDenied, ValidationError
from django.db import models
from collections import defaultdict
from decimal import Decimal

# Export-specific logger
export_logger = logging.getLogger('export.api')
backup_logger = logging.getLogger('backup')

try:
    from core.ml_integrations import ml_service, ML_AVAILABLE
except ImportError:
    ml_service = None
    ML_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

# Constants for performance and maintainability
class ExportConstants:
    MAX_RECORDS_PDF = 1000
    MAX_RECORDS_CSV = 5000
    CACHE_TIMEOUT = 300  # 5 minutes
    PDF_CHUNK_SIZE = 100
    
    # Column widths for PDF tables
    STUDENT_COL_WIDTHS = [0.8, 1.8, 1.0, 1.0, 1.0, 1.2, 0.8]
    
    # CSV Headers mapping (cached) - UPDATED: Removed fee columns
    CSV_HEADERS = {
        'students': ['Admission Number', 'First Name', 'Last Name', 'Father Name', 'Mother Name', 
                    'Date of Birth', 'Date of Admission', 'Gender', 'Religion', 'Caste Category',
                    'Mobile Number', 'Email', 'Address', 'Class Section', 'Aadhaar Number', 
                    'PEN Number', 'Blood Group', 'Attendance %', 'Final Due'],
        'teachers': ['Name', 'Mobile', 'Email', 'Qualification', 'Joining Date'],
        'subjects': ['Class Name', 'Section Name', 'Room Number', 'Student Count'],
        'transport': ['Route Name', 'Stoppage Name', 'Student Name', 'Assigned Date'],
        'fees': ['Fee Group', 'Group Type', 'Fee Type', 'Amount Type', 'Context Type', 'Month Name', 'Class Name', 'Stoppage Name', 'Amount'],
        'student_fees': ['Student', 'Receipt No', 'Amount', 'Discount', 'Paid Amount', 
                       'Deposit Date', 'Payment Mode', 'Transaction No', 'Payment Source', 'Note'],
        'fines': ['Fine Type', 'Category', 'Target Scope', 'Amount', 'Dynamic %', 'Reason', 'Due Date', 'Applied Date', 'Auto Generated', 'Created By', 'Class Section'],
        'attendance': ['Student', 'Class Section', 'Date', 'Status', 'Created At'],
        'promotion': ['Student', 'From Class', 'To Class', 'Academic Year', 'Promotion Date', 'Remarks', 'Min Percentage'],
        'messaging': ['Sender', 'Message Type', 'Recipient Type', 'Content', 'Total Recipients', 'Successful', 'Failed', 'Status', 'Source Module', 'Created At'],
        'users': ['Username', 'Email', 'Role', 'Mobile', 'Is Active', 'Is Staff', 'Is Superuser', 'Date Joined', 'Last Login'],
        'fees_report': ['Student Name', 'Admission Number', 'Class Section', 'Total Fees', 
                       'Amount Paid', 'Discount', 'CF Due', 'Fine Paid', 'Fine Unpaid', 
                       'Final Due', 'Payment Status', 'Mobile Number', 'Email']
    }
class DataExportService:
    """Centralized export service with enhanced security and performance"""
    
    SUPPORTED_MODULES = {
        'students': 'students.models.Student',
        'teachers': 'teachers.models.Teacher', 
        'subjects': 'subjects.models.ClassSection',
        'transport': 'transport.models.TransportAssignment',
        'fees': 'fees.models.FeesGroup',
        'student_fees': 'student_fees.models.FeeDeposit',
        'fines': 'fines.models.Fine',
        'attendance': 'attendance.models.Attendance',
        'promotion': 'promotion.models.StudentPromotion',
        'messaging': 'messaging.models.MessageLog',
        'users': 'users.models.CustomUser',
        'fees_report': 'reports.models.FeeReport'
    }
    
    @classmethod
    def _validate_user_permissions(cls, user, module_name):
        """Validate user permissions using server-side session data"""
        if not user.is_authenticated:
            raise PermissionDenied("Authentication required")
        
        # Superuser has all permissions
        if user.is_superuser:
            return True
        
        # Check module-specific permissions using the existing permission system
        from users.models import UserModulePermission
        permissions = UserModulePermission.get_user_permissions(user)
        
        # For export, user needs at least view permission on the module
        module_perms = permissions.get(module_name, {})
        if not module_perms.get('view', False):
            raise PermissionDenied(f"Export permission denied for {module_name} module")
        
        return True
    
    @classmethod
    def _sanitize_filename(cls, filename):
        """Sanitize filename to prevent path traversal"""
        # Remove any path components and dangerous characters
        import re
        filename = os.path.basename(filename)
        filename = re.sub(r'[^\w\-_\.]', '_', filename)
        return filename[:100]  # Limit length
    
    @classmethod
    def _get_cached_fee_breakdown(cls, student_ids):
        """Calculate fee breakdowns using reports module"""
        cache_key = f"fee_breakdown_bulk_{hash(tuple(sorted(student_ids)))}"
        cached_data = cache.get(cache_key)
        
        if cached_data:
            return cached_data
        
        # Direct database calculation (working logic from debug)
        try:
            from student_fees.models import FeeDeposit
            from fines.models import FineStudent
            from fees.models import FeesType
            from students.models import Student
            
            breakdowns = {}
            
            for student_id in student_ids:
                # Get total paid from fee deposits
                total_paid = FeeDeposit.objects.filter(
                    student_id=student_id
                ).aggregate(total=models.Sum('paid_amount'))['total'] or 0
                
                # Get fines through FineStudent relationship
                fine_students = FineStudent.objects.filter(
                    student_id=student_id, is_paid=False
                ).select_related('fine')
                fines_total = sum(fs.fine.amount or 0 for fs in fine_students)
                
                # Get student-specific fees based on their class
                try:
                    student = Student.objects.get(id=student_id)
                    if student.class_section:
                        # Get fees applicable to this student's class
                        current_fees = FeesType.objects.filter(
                            models.Q(class_name=student.class_section.class_name) |
                            models.Q(class_name__isnull=True) |
                            models.Q(class_name='')
                        ).aggregate(total=models.Sum('amount'))['total'] or 0
                    else:
                        current_fees = 0
                except Student.DoesNotExist:
                    current_fees = 0
                
                # Calculate CF due and total due
                cf_due = max(0, current_fees - total_paid)
                total_due = cf_due + fines_total
                
                breakdowns[student_id] = {
                    'cf_due': cf_due,
                    'current_fees': current_fees,
                    'total_paid': total_paid,
                    'fines': fines_total,
                    'total_due': total_due
                }
            
            cache.set(cache_key, breakdowns, ExportConstants.CACHE_TIMEOUT)
            return breakdowns
            
        except Exception as e:
            export_logger.error(f"Fee breakdown calculation failed: {e}")
            return {sid: cls._get_default_fee_breakdown() for sid in student_ids}
    
    @classmethod
    def _get_default_fee_breakdown(cls):
        """Default fee breakdown for error cases"""
        return {
            'cf_due': 0,
            'current_fees': 0,
            'total_paid': 0,
            'fines': 0,
            'total_due': 0
        }
    
    @classmethod
    def get_module_data(cls, module_name):
        """Get data using existing optimized managers"""
        start_time = time.time()
        export_logger.info(f"Fetching {module_name} data for export")
        
        try:
            if module_name == 'students':
                from students.models import Student
                data = Student.objects.get_list_optimized()
                return cls._organize_students_by_class(data)
            elif module_name == 'teachers':
                from teachers.models import Teacher
                return list(Teacher.objects.all())
            elif module_name == 'subjects':
                from subjects.models import ClassSection
                return list(ClassSection.objects.ordered())
            elif module_name == 'transport':
                from transport.models import TransportAssignment
                return list(TransportAssignment.objects.select_related('student', 'route', 'stoppage'))
            elif module_name == 'fees':
                from fees.models import FeesType
                return list(FeesType.objects.select_related('fee_group').all())
            elif module_name == 'student_fees':
                from student_fees.models import FeeDeposit
                return list(FeeDeposit.objects.select_related('student').order_by('-deposit_date'))
            elif module_name == 'fines':
                from fines.models import Fine
                return list(Fine.objects.select_related('fine_type', 'class_section', 'created_by').all())
            elif module_name == 'attendance':
                from attendance.models import Attendance
                return list(Attendance.objects.select_related('student', 'class_section'))
            elif module_name == 'promotion':
                from promotion.models import StudentPromotion
                return list(StudentPromotion.objects.select_related('student', 'from_class_section', 'to_class_section'))
            elif module_name == 'messaging':
                from messaging.models import MessageLog
                return list(MessageLog.objects.select_related('sender', 'class_section_filter').all())
            elif module_name == 'users':
                from users.models import CustomUser
                return list(CustomUser.objects.all().order_by('username'))
            elif module_name == 'fees_report':
                return cls._get_fee_report_data()
            return []
        except Exception as e:
            fetch_time = time.time() - start_time
            export_logger.error(f"Error fetching {module_name} data: {e}")
            return []
        finally:
            fetch_time = time.time() - start_time
            export_logger.info(f"Fetched {module_name} data in {fetch_time:.2f}s")
    
    @classmethod
    def _organize_students_by_class(cls, students):
        """Organize students by class for better export structure"""
        return sorted(students, key=lambda s: (s.class_section.class_name if s.class_section else 'No Class', s.first_name))
    

    
    @classmethod
    def _write_students_csv(cls, writer, students):
        """Write student data to CSV with Final Due from fee report API"""
        # Get fee report data for Final Due calculations using centralized method
        fee_report_data = cls._get_fee_report_data()
        fee_lookup = {row.student_id: row.final_due for row in fee_report_data}
        
        for student in students:
            # Get Final Due from centralized fee report calculation
            final_due = fee_lookup.get(student.id, 0)
            
            writer.writerow([
                student.admission_number or '',
                student.first_name or '',
                student.last_name or '',
                student.father_name or '',
                student.mother_name or '',
                student.date_of_birth.strftime('%Y-%m-%d') if student.date_of_birth else '',
                student.date_of_admission.strftime('%Y-%m-%d') if student.date_of_admission else '',
                student.gender or '',
                student.religion or '',
                student.caste_category or '',
                student.mobile_number or '',
                student.email or '',
                student.address or '',
                str(student.class_section) if student.class_section else '',
                student.aadhaar_number or '',
                student.pen_number or '',
                student.blood_group or '',
                '85%',  # Placeholder for attendance
                f"{float(final_due):.2f}"  # Final Due from centralized fee report API
            ])
    
    @classmethod
    def _write_teachers_csv(cls, writer, teachers):
        for teacher in teachers:
            writer.writerow([
                f"{teacher.first_name} {teacher.last_name}",
                teacher.mobile_number or '',
                teacher.email or '',
                teacher.qualification or '',
                teacher.joining_date.strftime('%Y-%m-%d') if teacher.joining_date else ''
            ])
    
    @classmethod
    def _write_subjects_csv(cls, writer, class_sections):
        for cs in class_sections:
            writer.writerow([
                cs.class_name or '',
                cs.section_name or '',
                cs.room_number or '',
                cs.student_count or 0
            ])
    
    @classmethod
    def _write_transport_csv(cls, writer, assignments):
        for assignment in assignments:
            writer.writerow([
                assignment.route.route_name if assignment.route else '',
                assignment.stoppage.stoppage_name if assignment.stoppage else '',
                f"{assignment.student.first_name} {assignment.student.last_name}" if assignment.student else '',
                assignment.assigned_date.strftime('%Y-%m-%d') if assignment.assigned_date else ''
            ])
    
    @classmethod
    def _write_fees_csv(cls, writer, fee_types):
        for fee_type in fee_types:
            writer.writerow([
                fee_type.fee_group.group_name if fee_type.fee_group else '',
                fee_type.fee_group.group_type if fee_type.fee_group else '',
                fee_type.fee_type or '',
                fee_type.amount_type or '',
                fee_type.context_type or '',
                fee_type.month_name or '',
                fee_type.class_name or '',
                fee_type.stoppage_name or '',
                fee_type.amount or 0
            ])
    
    @classmethod
    def _write_student_fees_csv(cls, writer, deposits):
        for deposit in deposits:
            writer.writerow([
                f"{deposit.student.first_name} {deposit.student.last_name}" if deposit.student else '',
                deposit.receipt_no or '',
                deposit.amount or 0,
                deposit.discount or 0,
                deposit.paid_amount or 0,
                deposit.deposit_date.strftime('%Y-%m-%d') if deposit.deposit_date else '',
                deposit.payment_mode or '',
                deposit.transaction_no or '',
                deposit.payment_source or '',
                deposit.note or ''
            ])
    
    @classmethod
    def _write_fines_csv(cls, writer, fines):
        for fine in fines:
            writer.writerow([
                fine.fine_type.fine_type if fine.fine_type else '',
                fine.fine_type.category if fine.fine_type else '',
                fine.fine_type.target_scope if fine.fine_type else '',
                fine.amount or 0,
                fine.fine_type.dynamic_percentage if fine.fine_type else 0,
                fine.reason or '',
                fine.due_date.strftime('%Y-%m-%d') if fine.due_date else '',
                fine.applied_date.strftime('%Y-%m-%d') if fine.applied_date else '',
                'Yes' if fine.auto_generated else 'No',
                fine.created_by.username if fine.created_by else '',
                str(fine.class_section) if fine.class_section else ''
            ])
    
    @classmethod
    def _write_attendance_csv(cls, writer, attendance_records):
        for record in attendance_records:
            writer.writerow([
                f"{record.student.first_name} {record.student.last_name}" if record.student else '',
                str(record.class_section) if record.class_section else '',
                record.date.strftime('%Y-%m-%d') if record.date else '',
                'Present' if record.status else 'Absent',
                record.created_at.strftime('%Y-%m-%d %H:%M') if record.created_at else ''
            ])
    
    @classmethod
    def _write_promotion_csv(cls, writer, promotions):
        for promotion in promotions:
            writer.writerow([
                f"{promotion.student.first_name} {promotion.student.last_name}" if promotion.student else '',
                str(promotion.from_class_section) if promotion.from_class_section else '',
                str(promotion.to_class_section) if promotion.to_class_section else '',
                promotion.academic_year or '',
                promotion.promotion_date.strftime('%Y-%m-%d') if promotion.promotion_date else '',
                promotion.remarks or '',
                promotion.min_percentage or 0
            ])
    
    @classmethod
    def _write_messaging_csv(cls, writer, messages):
        for message in messages:
            writer.writerow([
                message.sender.username if message.sender else '',
                message.message_type or '',
                message.recipient_type or '',
                message.content[:100] + '...' if len(message.content or '') > 100 else message.content or '',
                message.total_recipients or 0,
                message.successful_count or 0,
                message.failed_count or 0,
                message.status or '',
                message.source_module or '',
                message.created_at.strftime('%Y-%m-%d %H:%M') if message.created_at else ''
            ])
    
    @classmethod
    def _write_users_csv(cls, writer, users):
        for user in users:
            writer.writerow([
                user.username or '',
                user.email or '',
                user.role or '',
                user.mobile or '',
                'Yes' if user.is_active else 'No',
                'Yes' if user.is_staff else 'No',
                'Yes' if user.is_superuser else 'No',
                user.date_joined.strftime('%Y-%m-%d') if user.date_joined else '',
                user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else ''
            ])
    
    @classmethod
    def _organize_students_by_class(cls, students):
        """Organize students by class from Nursery to Class 10"""
        from core.utils import extract_class_number
        
        # Group students by class
        class_groups = defaultdict(list)
        for student in students:
            try:
                if hasattr(student, 'class_section') and student.class_section:
                    class_name = getattr(student.class_section, 'class_name', 'Unknown')
                    class_groups[class_name].append(student)
                else:
                    class_groups['Unassigned'].append(student)
            except AttributeError as e:
                export_logger.warning(f"Error organizing student: {e}")
                class_groups['Unassigned'].append(student)
        
        # Sort classes in educational order
        try:
            sorted_classes = sorted(class_groups.keys(), key=extract_class_number)
        except ImportError:
            sorted_classes = sorted(class_groups.keys())
        
        # Flatten back to list with class order maintained
        organized_students = []
        for class_name in sorted_classes:
            # Sort students within each class by name safely
            try:
                class_students = sorted(class_groups[class_name], key=lambda s: (
                    getattr(s, 'first_name', ''), 
                    getattr(s, 'last_name', '')
                ))
            except AttributeError:
                class_students = class_groups[class_name]
            organized_students.extend(class_students)
        
        return organized_students
    
    @classmethod
    def _format_due_amount(cls, amount):
        """Format due amount with safe conversion"""
        try:
            if amount is None or amount == '':
                return '0.00'
            
            # Handle different numeric types safely
            if isinstance(amount, (int, float)):
                decimal_amount = Decimal(str(amount))
            else:
                # Try to convert string to Decimal
                decimal_amount = Decimal(str(amount))
            
            return f'{decimal_amount:,.2f}'
            
        except (ValueError, TypeError, AttributeError):
            return '0.00'
    
    @classmethod
    def _get_detailed_fee_breakdown(cls, student):
        """Get detailed fee breakdown for export display (deprecated - use bulk method)"""
        export_logger.warning("Using deprecated single student fee breakdown - consider bulk calculation")
        
        try:
            student_id = getattr(student, 'id', None)
            if not student_id:
                return cls._get_default_fee_breakdown()
            
            breakdowns = cls._get_cached_fee_breakdown([student_id])
            return breakdowns.get(student_id, cls._get_default_fee_breakdown())
            
        except Exception as e:
            export_logger.error(f"Fee breakdown calculation failed for student: {e}")
            return cls._get_default_fee_breakdown()
    
    @classmethod
    def export_to_csv(cls, module_name, user):
        """Export module data to CSV format with enhanced security"""
        start_time = time.time()
        
        try:
            # Enhanced security validation
            cls._validate_user_permissions(user, module_name)
            
            data = cls.get_module_data(module_name)
            
            # Validate data size
            if len(data) > ExportConstants.MAX_RECORDS_CSV:
                raise ValidationError(f"Dataset too large: {len(data)} records. Maximum allowed: {ExportConstants.MAX_RECORDS_CSV}")
            
            export_logger.info(f"CSV Export Started: {module_name} - {len(data)} records - User: {user.username}")
            
            return cls._generate_csv_response(module_name, data, user, start_time)
            
        except (PermissionDenied, ValidationError) as e:
            export_logger.warning(f"CSV export denied for {module_name}: {e} - User: {user.username}")
            return HttpResponse(str(e), status=403)
        except Exception as e:
            export_logger.error(f"CSV export failed for {module_name}: {e}")
            return HttpResponse(f'Export failed: {str(e)}', status=500)
    
    @classmethod
    def _generate_csv_response(cls, module_name, data, user, start_time):
        """Generate CSV response with proper error handling"""
        try:
            # Create CSV response with sanitized filename
            current_time = datetime.now()
            timestamp = current_time.strftime('%Y%m%d_%H%M')
            filename = cls._sanitize_filename(f"{module_name}_export_{timestamp}.csv")
            
            response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            writer = csv.writer(response)
            
            # Write metadata header
            writer.writerow([f'Exported on {current_time.strftime("%Y-%m-%d %H:%M:%S")} by {user.username}, Total Records: {len(data)}'])
            writer.writerow([])  # Empty row
            
            # Write headers from cached mapping
            headers = ExportConstants.CSV_HEADERS.get(module_name, ['Data'])
            writer.writerow(headers)
            
            # Pre-load fee report data for students to use centralized calculation
            if module_name == 'students' and data:
                fee_report_data = cls._get_fee_report_data()
                cls._fee_report_cache = {row.student_id: row.final_due for row in fee_report_data}
            
            # Write data rows with enhanced error handling
            processed_count = 0
            for i, item in enumerate(data):
                try:
                    row_data = cls._get_csv_row_safe(module_name, item)
                    writer.writerow(row_data)
                    processed_count += 1
                except AttributeError as e:
                    export_logger.warning(f"Missing attribute in row {i} for {module_name}: {e}")
                    continue
                except Exception as e:
                    export_logger.error(f"Error processing row {i} for {module_name}: {e}")
                    continue
            
            # Clean up cache
            if hasattr(cls, '_fee_report_cache'):
                delattr(cls, '_fee_report_cache')
            
            total_time = time.time() - start_time
            export_logger.info(
                f"CSV Export Complete: {module_name} - {filename} - "
                f"{processed_count}/{len(data)} records - {total_time:.3f}s - User: {user.username}"
            )
            
            return response
            
        except Exception as e:
            export_logger.error(f"CSV generation failed for {module_name}: {e}")
            return HttpResponse(f'Export failed: {str(e)}', status=500)
    
    @classmethod
    def _get_csv_row_safe(cls, module_name, item, fee_breakdowns=None):
        """Get CSV row data with safe attribute access and error handling"""
        def safe_get(obj, attr, default='N/A'):
            """Safely get attribute with default value"""
            try:
                value = getattr(obj, attr, default)
                return str(value) if value is not None else default
            except AttributeError:
                return default
        
        def safe_format_date(date_obj, format_str='%Y-%m-%d'):
            """Safely format date object"""
            try:
                return date_obj.strftime(format_str) if date_obj else 'N/A'
            except (AttributeError, ValueError):
                return 'N/A'
        
        try:
            if module_name == 'students':
                # Get final due from centralized fee report calculation
                final_due = 0
                try:
                    # Use centralized fee report data for consistent calculations
                    student_id = getattr(item, 'id', None)
                    if student_id and hasattr(cls, '_fee_report_cache'):
                        # Use cached fee report data if available
                        final_due = cls._fee_report_cache.get(student_id, 0)
                    elif student_id:
                        # Get from fee report data using same method as Student Fee Report
                        fee_report_data = cls._get_fee_report_data()
                        for fee_row in fee_report_data:
                            if getattr(fee_row, 'student_id', None) == student_id:
                                final_due = getattr(fee_row, 'final_due', 0)
                                break
                except Exception:
                    final_due = 0
                
                return [
                    safe_get(item, 'admission_number'),
                    safe_get(item, 'first_name'),
                    safe_get(item, 'last_name'),
                    safe_get(item, 'father_name'),
                    safe_get(item, 'mother_name'),
                    safe_format_date(getattr(item, 'date_of_birth', None)),
                    safe_format_date(getattr(item, 'date_of_admission', None)),
                    safe_get(item, 'gender'),
                    safe_get(item, 'religion'),
                    safe_get(item, 'caste_category'),
                    safe_get(item, 'mobile_number'),
                    safe_get(item, 'email'),
                    safe_get(item, 'address'),
                    str(item.class_section) if hasattr(item, 'class_section') and item.class_section else 'N/A',
                    safe_get(item, 'aadhaar_number'),
                    safe_get(item, 'pen_number'),
                    safe_get(item, 'blood_group'),
                    f"{getattr(item, 'attendance_percentage', 0):.1f}%",
                    f"{float(final_due):.2f}"
                ]
            elif module_name == 'teachers':
                return [
                    safe_get(item, 'name'),
                    safe_get(item, 'mobile'),
                    safe_get(item, 'email'),
                    safe_get(item, 'qualification'),
                    safe_format_date(getattr(item, 'joining_date', None))
                ]
            elif module_name == 'subjects':
                # Avoid N+1 query by using cached count or safe access
                student_count = 'N/A'
                try:
                    if hasattr(item, '_prefetched_objects_cache') and 'students' in item._prefetched_objects_cache:
                        student_count = len(item._prefetched_objects_cache['students'])
                    elif hasattr(item, 'student_count'):
                        student_count = item.student_count
                except Exception:
                    student_count = 'N/A'
                
                return [
                    safe_get(item, 'class_name'),
                    safe_get(item, 'section_name'),
                    safe_get(item, 'room_number'),
                    student_count
                ]
            elif module_name == 'fees':
                return [
                    str(item.fee_group) if hasattr(item, 'fee_group') else 'N/A',
                    safe_get(item.fee_group, 'group_type') if hasattr(item, 'fee_group') else 'N/A',
                    safe_get(item.fee_group, 'fee_type') if hasattr(item, 'fee_group') else 'N/A',
                    safe_get(item, 'amount_type'),
                    safe_get(item, 'context_type'),
                    safe_get(item, 'month_name'),
                    safe_get(item, 'class_name'),
                    safe_get(item, 'stoppage_name'),
                    safe_get(item, 'amount')
                ]
            elif module_name == 'student_fees':
                return [
                    str(item.student) if hasattr(item, 'student') else 'N/A',
                    safe_get(item, 'receipt_no'),
                    safe_get(item, 'amount'),
                    safe_get(item, 'discount'),
                    safe_get(item, 'paid_amount'),
                    safe_format_date(getattr(item, 'deposit_date', None)),
                    safe_get(item, 'payment_mode'),
                    safe_get(item, 'transaction_no'),
                    safe_get(item, 'payment_source'),
                    safe_get(item, 'note')
                ]
            elif module_name == 'fines':
                return [
                    str(item.fine_type) if hasattr(item, 'fine_type') else 'N/A',
                    safe_get(item.fine_type, 'category') if hasattr(item, 'fine_type') else 'N/A',
                    safe_get(item, 'target_scope'),
                    safe_get(item, 'amount'),
                    safe_get(item, 'dynamic_amount_percent'),
                    safe_get(item, 'reason'),
                    safe_format_date(getattr(item, 'due_date', None)),
                    safe_format_date(getattr(item, 'applied_date', None)),
                    'Yes' if getattr(item, 'auto_generated', False) else 'No',
                    str(item.created_by) if hasattr(item, 'created_by') and item.created_by else 'N/A',
                    str(item.class_section) if hasattr(item, 'class_section') and item.class_section else 'N/A'
                ]
            elif module_name == 'attendance':
                return [
                    str(item.student) if hasattr(item, 'student') else 'N/A',
                    str(item.class_section) if hasattr(item, 'class_section') and item.class_section else 'N/A',
                    safe_format_date(getattr(item, 'date', None)),
                    safe_get(item, 'status'),
                    safe_format_date(getattr(item, 'created_at', None), '%Y-%m-%d %H:%M:%S')
                ]
            elif module_name == 'promotion':
                return [
                    str(item.student) if hasattr(item, 'student') else 'N/A',
                    str(item.from_class_section) if hasattr(item, 'from_class_section') and item.from_class_section else 'N/A',
                    str(item.to_class_section) if hasattr(item, 'to_class_section') and item.to_class_section else 'N/A',
                    safe_get(item, 'academic_year'),
                    safe_format_date(getattr(item, 'promotion_date', None)),
                    safe_get(item, 'remarks'),
                    safe_get(item, 'minimum_percentage', 'N/A')
                ]
            elif module_name == 'messaging':
                return [
                    str(item.sender) if hasattr(item, 'sender') and item.sender else 'N/A',
                    safe_get(item, 'message_type'),
                    safe_get(item, 'recipient_type'),
                    safe_get(item, 'message_content')[:100] + '...' if len(safe_get(item, 'message_content', '')) > 100 else safe_get(item, 'message_content'),
                    safe_get(item, 'total_recipients'),
                    safe_get(item, 'successful_sends'),
                    safe_get(item, 'failed_sends'),
                    safe_get(item, 'status'),
                    safe_get(item, 'source_module'),
                    safe_format_date(getattr(item, 'created_at', None), '%Y-%m-%d %H:%M:%S')
                ]
            elif module_name == 'users':
                return [
                    safe_get(item, 'username'),
                    safe_get(item, 'email'),
                    safe_get(item, 'role'),
                    safe_get(item, 'mobile'),
                    'Yes' if getattr(item, 'is_active', False) else 'No',
                    'Yes' if getattr(item, 'is_staff', False) else 'No',
                    'Yes' if getattr(item, 'is_superuser', False) else 'No',
                    safe_format_date(getattr(item, 'date_joined', None), '%Y-%m-%d %H:%M:%S'),
                    safe_format_date(getattr(item, 'last_login', None), '%Y-%m-%d %H:%M:%S')
                ]
            elif module_name == 'fees_report':
                # Use dynamic values from fee report calculation
                return [
                    safe_get(item, 'name'),
                    safe_get(item, 'admission_number'),
                    safe_get(item, 'class_name'),
                    f"{getattr(item, 'current_fees', 0):.2f}",
                    f"{getattr(item, 'current_paid', 0):.2f}",
                    f"{getattr(item, 'current_discount', 0):.2f}",
                    f"{getattr(item, 'cf_due', 0):.2f}",
                    f"{getattr(item, 'fine_paid', 0):.2f}",
                    f"{getattr(item, 'fine_unpaid', 0):.2f}",
                    f"{getattr(item, 'final_due', 0):.2f}",
                    'Paid' if getattr(item, 'final_due', 0) == 0 else 'Outstanding',
                    safe_get(item, 'mobile_number'),
                    safe_get(item, 'email')
                ]
            else:
                return [str(item)]
                
        except Exception as e:
            export_logger.error(f"Error processing CSV row for {module_name}: {e}")
            return ['Error processing data']
    
    @classmethod
    def export_to_pdf(cls, module_name, user):
        """Export data to PDF format with enhanced security and performance"""
        start_time = time.time()
        
        try:
            # Enhanced security validation
            cls._validate_user_permissions(user, module_name)
            
            # Validate data size for PDF performance
            data = cls.get_module_data(module_name)
            if len(data) > ExportConstants.MAX_RECORDS_PDF:
                raise ValidationError(f"Dataset too large for PDF: {len(data)} records. Maximum allowed: {ExportConstants.MAX_RECORDS_PDF}")
            
            export_logger.info(f"PDF Export Started: {module_name} - {len(data)} records - User: {user.username}")
            
            return cls._generate_pdf_response(module_name, data, user, start_time)
            
        except (PermissionDenied, ValidationError) as e:
            export_logger.warning(f"PDF export denied for {module_name}: {e} - User: {user.username}")
            return HttpResponse(str(e), status=403)
        except Exception as e:
            export_logger.error(f"PDF export failed for {module_name}: {e}")
            return HttpResponse(f'PDF export failed: {str(e)}', status=500)
    
    @classmethod
    def _setup_pdf_document(cls, buffer):
        """Setup PDF document with proper configuration for wide tables"""
        return SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=0.3*inch,
            leftMargin=0.3*inch,
            topMargin=0.5*inch,
            bottomMargin=0.3*inch
        )
    
    @classmethod
    def _create_pdf_styles(cls):
        """Create PDF styles"""
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=20,
            alignment=1,
            textColor=colors.darkblue
        )
        return styles, title_style
    
    @classmethod
    def _generate_pdf_response(cls, module_name, data, user, start_time):
        """Generate PDF response with proper resource management"""
        try:
            if not PDF_AVAILABLE:
                export_logger.warning(f"PDF library unavailable, falling back to CSV for {module_name}")
                return cls.export_to_csv(module_name, user)
            
            # Create PDF buffer
            buffer = io.BytesIO()
            doc = cls._setup_pdf_document(buffer)
            
            # Container for PDF elements
            elements = []
            styles, title_style = cls._create_pdf_styles()
            
            # Generate PDF content
            cls._generate_pdf_content(elements, module_name, data, user, styles, title_style)
            
            # Build PDF with error handling
            generation_start = time.time()
            doc.build(elements)
            generation_time = time.time() - generation_start
            export_logger.debug(f"PDF generation completed in {generation_time:.3f}s")
            
            # Create response BEFORE closing buffer
            response = cls._create_pdf_response(buffer, module_name, user, start_time)
            
            # Close buffer after response creation
            buffer.close()
            
            return response
            
        except Exception as e:
            export_logger.error(f"PDF generation error for {module_name}: {e}")
            return HttpResponse(f'PDF export failed: {str(e)}', status=500)
    
    @classmethod
    def _generate_pdf_content(cls, elements, module_name, data, user, styles, title_style):
        """Generate PDF content elements"""
        current_time = datetime.now()
        
        # Title
        title = Paragraph(f"{module_name.title()} Export Report", title_style)
        elements.append(title)
        
        # Metadata
        meta_style = ParagraphStyle(
            'MetaStyle',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=15
        )
        meta_text = f"Generated on {current_time.strftime('%Y-%m-%d %H:%M:%S')} by {user.username} | Total Records: {len(data)}"
        elements.append(Paragraph(meta_text, meta_style))
        elements.append(Spacer(1, 12))
        
        if module_name == 'students' and data:
            cls._create_students_pdf_by_class(elements, data)
        else:
            cls._create_standard_pdf_table(elements, module_name, data)
    
    @classmethod
    def _create_pdf_response(cls, buffer, module_name, user, start_time):
        """Create HTTP response for PDF"""
        try:
            # Get PDF content while buffer is still open
            pdf_content = buffer.getvalue()
            
            # Validate PDF content
            if not pdf_content or len(pdf_content) < 100:
                raise ValueError(f"PDF validation failed: {len(pdf_content)} bytes")
            
            export_logger.info(f"PDF content validated: {len(pdf_content)} bytes for {module_name}")
            
            # Create sanitized filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M')
            filename = cls._sanitize_filename(f"{module_name}_export_{timestamp}.pdf")
            
            response = HttpResponse(pdf_content, content_type='application/pdf')
            response['Content-Disposition'] = f'inline; filename="{filename}"'
            response['Content-Length'] = len(pdf_content)
            response['Accept-Ranges'] = 'bytes'
            response['X-Content-Type-Options'] = 'nosniff'
            response['Cache-Control'] = 'public, max-age=300'
            response['X-PDF-Viewer'] = 'chrome-native'
            
            total_time = time.time() - start_time
            export_logger.info(
                f"PDF Export Complete: {module_name} - {filename} - "
                f"{len(pdf_content)} bytes - {total_time:.3f}s - User: {user.username}"
            )
            
            return response
            
        except Exception as e:
            export_logger.error(f"PDF response creation failed: {e}")
            raise
    
    @classmethod
    def _create_students_pdf_by_class(cls, elements, students):
        """Create PDF with students organized by class - refactored for maintainability"""
        class_groups = cls._group_students_by_class(students)
        sorted_classes = cls._sort_classes(class_groups.keys())
        
        for class_name in sorted_classes:
            if not class_groups[class_name]:
                continue
            
            cls._create_class_section(elements, class_name, class_groups[class_name])
            
            # Page break after each class (except last)
            if class_name != sorted_classes[-1]:
                elements.append(PageBreak())
    
    @classmethod
    def _group_students_by_class(cls, students):
        """Group students by class with safe attribute access"""
        class_groups = defaultdict(list)
        for student in students:
            try:
                if hasattr(student, 'class_section') and student.class_section:
                    class_name = getattr(student.class_section, 'class_name', 'Unknown')
                    class_groups[class_name].append(student)
                else:
                    class_groups['Unassigned'].append(student)
            except AttributeError as e:
                export_logger.warning(f"Error grouping student: {e}")
                class_groups['Unassigned'].append(student)
        
        return class_groups
    
    @classmethod
    def _sort_classes(cls, class_names):
        """Sort class names safely"""
        try:
            from core.utils import extract_class_number
            return sorted(class_names, key=extract_class_number)
        except ImportError:
            return sorted(class_names)
    
    @classmethod
    def _create_class_section(cls, elements, class_name, students):
        """Create a class section in PDF"""
        # Class header
        styles = getSampleStyleSheet()
        class_style = ParagraphStyle(
            'ClassHeader',
            parent=styles['Heading2'],
            fontSize=12,
            spaceAfter=10,
            textColor=colors.darkblue
        )
        
        class_header = Paragraph(f"{class_name} ({len(students)} students)", class_style)
        elements.append(class_header)
        
        # Create table
        table = cls._create_students_table(students)
        elements.append(table)
        elements.append(Spacer(1, 15))
    
    @classmethod
    def _create_students_table(cls, students):
        """Create students table with bulk fee calculation"""
        # Headers
        compact_headers = [
            'Adm. No.', 'Name', 'Father Name', 'Mother Name', 'Class', 'Mobile', 'Final Due'
        ]
        
        table_data = [compact_headers]
        
        # Get fee report data for Final Due calculations using centralized method
        fee_report_data = cls._get_fee_report_data()
        fee_lookup = {row.student_id: row.final_due for row in fee_report_data}
        
        # Sort students safely
        try:
            sorted_students = sorted(students, key=lambda s: (
                getattr(s, 'first_name', ''), 
                getattr(s, 'last_name', '')
            ))
        except AttributeError:
            sorted_students = students
        
        for student in sorted_students:
            row = cls._create_student_row(student, fee_lookup)
            table_data.append(row)
        
        # Create table with constants
        col_widths = [w*inch for w in ExportConstants.STUDENT_COL_WIDTHS]
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(cls._get_table_style())
        
        return table
    
    @classmethod
    def _create_student_row(cls, student, fee_lookup):
        """Create a single student row with safe attribute access"""
        def safe_slice(value, length):
            """Safely slice string value"""
            try:
                return str(value or '')[:length]
            except (AttributeError, TypeError):
                return 'N/A'
        
        student_id = getattr(student, 'id', None)
        final_due = fee_lookup.get(student_id, 0) if student_id else 0
        
        return [
            safe_slice(getattr(student, 'admission_number', ''), 8),
            f"{getattr(student, 'first_name', '')} {getattr(student, 'last_name', '')}"[:15],
            safe_slice(getattr(student, 'father_name', ''), 12),
            safe_slice(getattr(student, 'mother_name', ''), 12),
            safe_slice(str(student.class_section) if hasattr(student, 'class_section') and student.class_section else 'N/A', 8),
            safe_slice(getattr(student, 'mobile_number', ''), 10),
            f"{float(final_due):.0f}"
        ]
    
    @classmethod
    def _get_table_style(cls):
        """Get table style configuration"""
        return TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            
            # Data styling
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            
            # Mobile and email alignment
            ('ALIGN', (3, 1), (4, -1), 'CENTER'),
        ])
    
    @classmethod
    def _create_standard_pdf_table(cls, elements, module_name, data):
        """Create standard PDF table with proper column widths"""
        headers = ExportConstants.CSV_HEADERS.get(module_name, ['Data'])
        
        # Process data in chunks for better performance
        limited_data = data[:ExportConstants.PDF_CHUNK_SIZE] if len(data) > ExportConstants.PDF_CHUNK_SIZE else data
        
        table_data = [headers]
        
        for item in limited_data:
            try:
                row_data = cls._get_csv_row_safe(module_name, item)
                # Truncate long text for PDF display with specific limits per column
                truncated_row = []
                for i, cell in enumerate(row_data):
                    if i == 0:  # Name column
                        truncated_row.append(str(cell)[:15] + '...' if len(str(cell)) > 15 else str(cell))
                    elif i in [1, 2]:  # Admission, Class
                        truncated_row.append(str(cell)[:10] + '...' if len(str(cell)) > 10 else str(cell))
                    elif i in [3, 4, 5, 6, 7, 8, 9]:  # Numeric columns
                        truncated_row.append(str(cell)[:8] + '...' if len(str(cell)) > 8 else str(cell))
                    else:  # Other columns
                        truncated_row.append(str(cell)[:12] + '...' if len(str(cell)) > 12 else str(cell))
                table_data.append(truncated_row)
            except Exception as e:
                export_logger.warning(f"Error processing item for PDF: {e}")
                continue
        
        # Custom column widths for different modules
        if module_name == 'fees_report':
            col_widths = [1.2*inch, 0.8*inch, 0.8*inch, 0.6*inch, 0.6*inch, 0.6*inch, 0.6*inch, 0.6*inch, 0.6*inch, 0.6*inch, 0.8*inch, 0.8*inch, 1.0*inch]
        elif module_name == 'students':
            # Custom widths for student export with 19 columns
            col_widths = [0.6*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.6*inch, 0.6*inch, 0.4*inch, 0.5*inch, 0.5*inch, 0.7*inch, 0.8*inch, 0.8*inch, 0.6*inch, 0.6*inch, 0.5*inch, 0.4*inch, 0.5*inch, 0.6*inch]
        else:
            # Dynamic column widths for other modules
            num_cols = len(headers)
            available_width = 10 * inch
            col_width = available_width / num_cols
            col_widths = [col_width] * num_cols
        
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 6),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 4),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 5),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('WORDWRAP', (0, 0), (-1, -1), True),
        ]))
        
        elements.append(table)
        
        # Add note if data was limited
        if len(data) > ExportConstants.PDF_CHUNK_SIZE:
            styles = getSampleStyleSheet()
            note = Paragraph(
                f"Note: Showing first {ExportConstants.PDF_CHUNK_SIZE} of {len(data)} records for PDF optimization.", 
                styles['Normal']
            )
            elements.append(Spacer(1, 10))
            elements.append(note)
    
    @classmethod
    def export_to_excel(cls, module_name, user):
        """Export module data to Excel format with enhanced formatting"""
        start_time = time.time()
        export_logger.info(f"Excel export started for {module_name} by {user.username}")
        
        try:
            # Enhanced security validation
            cls._validate_user_permissions(user, module_name)
            
            if not EXCEL_AVAILABLE:
                return cls.export_to_csv(module_name, user)
            
            # Use cached data for performance
            cache_key = f"excel_export_{module_name}_{user.id}"
            data = cache.get(cache_key)
            if not data:
                data = cls.get_module_data(module_name)
                cache.set(cache_key, data, 60)  # Cache for 1 minute
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = cls._sanitize_filename(f"{module_name}_export_{timestamp}.xlsx")
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb = Workbook()
            
            # Use single sheet format for all modules
            ws = wb.active
            ws.title = f"{module_name.title()} Export"
            
            # Add headers with formatting
            headers = ExportConstants.CSV_HEADERS.get(module_name, ['Data'])
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Pre-load fee report data for students to use centralized calculation
            if module_name == 'students' and data:
                fee_report_data = cls._get_fee_report_data()
                cls._fee_report_cache = {row.student_id: row.final_due for row in fee_report_data}
            
            # Add all data rows
            current_row = 2
            for item in data:
                try:
                    row_data = cls._get_csv_row_safe(module_name, item)
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=current_row, column=col_idx, value=value)
                    current_row += 1
                except Exception as e:
                    export_logger.warning(f"Error processing row in Excel export: {e}")
                    continue
            
            # Clean up cache
            if hasattr(cls, '_fee_report_cache'):
                delattr(cls, '_fee_report_cache')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(response)
            
            total_time = time.time() - start_time
            export_logger.info(f"Excel export completed for {module_name} in {total_time:.2f}s")
            
            return response
            
        except (PermissionDenied, ValidationError) as e:
            export_logger.warning(f"Excel export denied for {module_name}: {e}")
            return HttpResponse(str(e), status=403)
        except Exception as e:
            export_logger.error(f"Excel export failed for {module_name}: {e}")
            return HttpResponse(f'Excel export failed: {str(e)}', status=500)
    

    
    @classmethod
    def _create_fees_report_excel_by_class(cls, wb, report_data):
        """Create Excel with all fee report data in single sheet"""
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create single sheet with all students
        ws = wb.create_sheet("Fee Report")
        
        # Headers
        headers = ExportConstants.CSV_HEADERS.get('fees_report', ['Data'])
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add all students data
        current_row = 2
        for student in report_data:
            row_data = cls._get_csv_row_safe('fees_report', student)
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=current_row, column=col_idx, value=value)
            current_row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width

    @classmethod
    def _get_fee_report_data(cls):
        """Get fee report data using existing reports view logic"""
        from decimal import Decimal
        from django.db.models import Sum, Q
        from students.models import Student
        from student_fees.models import FeeDeposit
        from fines.models import FineStudent
        from fees.models import FeesType
        from transport.models import TransportAssignment
        
        students = Student.objects.select_related('class_section').all()
        raw_data = []
        
        for student in students:
            # Use same logic as fees_report view
            class_name = student.class_section.class_name if student.class_section else ''
            class_display = student.class_section.display_name if student.class_section else ''
            
            # Get applicable current session fees
            applicable_fees = FeesType.objects.filter(
                Q(class_name__isnull=True) | 
                Q(class_name__iexact=class_name) |
                Q(class_name__iexact=class_display)
            ).exclude(fee_group__group_type="Transport")
            
            # Add transport fees if assigned
            transport_assignment = TransportAssignment.objects.filter(student=student).first()
            if transport_assignment and transport_assignment.stoppage:
                transport_fees = FeesType.objects.filter(
                    fee_group__group_type="Transport",
                    related_stoppage=transport_assignment.stoppage
                )
                applicable_fees = list(applicable_fees) + list(transport_fees)
            
            # Calculate current session fees
            current_fees = sum(fee.amount for fee in applicable_fees)
            
            # Carry Forward amounts
            cf_due_original = student.due_amount or Decimal('0')
            
            # Total fees = Current session fees + Carry forward due
            total_fees = current_fees + cf_due_original
            
            # Get ALL payments (fees + CF, excluding fines)
            fee_payments = FeeDeposit.objects.filter(
                student=student
            ).exclude(note__icontains="Fine Payment")
            
            total_fee_paid = fee_payments.aggregate(Sum('paid_amount'))['paid_amount__sum'] or Decimal('0')
            total_fee_discount = fee_payments.aggregate(Sum('discount'))['discount__sum'] or Decimal('0')
            
            # Separate CF payments for display
            cf_payments = fee_payments.filter(
                Q(note__icontains="Carry Forward") | Q(payment_source="carry_forward")
            )
            cf_paid = cf_payments.aggregate(Sum('paid_amount'))['paid_amount__sum'] or Decimal('0')
            cf_discount = cf_payments.aggregate(Sum('discount'))['discount__sum'] or Decimal('0')
            
            # Remaining CF due
            cf_due = max(cf_due_original - cf_paid - cf_discount, Decimal('0'))
            
            # Fine amounts
            unpaid_fines = FineStudent.objects.filter(
                student=student, 
                is_paid=False
            ).select_related('fine')
            
            paid_fines = FineStudent.objects.filter(
                student=student, 
                is_paid=True
            ).select_related('fine')
            
            fine_unpaid = sum(fs.fine.amount for fs in unpaid_fines)
            fine_paid = sum(fs.fine.amount for fs in paid_fines)
            
            # Calculate final due: Total fees - Total paid - Total discount + Unpaid fines
            final_due = max(total_fees - total_fee_paid - total_fee_discount + fine_unpaid, Decimal('0'))
            
            row = type('FeeReportRow', (), {
                'name': f"{student.first_name} {student.last_name}",
                'admission_number': student.admission_number or '',
                'class_name': str(student.class_section) if student.class_section else 'Unassigned',
                'current_fees': float(total_fees),
                'current_paid': float(total_fee_paid),
                'current_discount': float(total_fee_discount),
                'cf_due': float(cf_due),
                'fine_paid': float(fine_paid),
                'fine_unpaid': float(fine_unpaid),
                'final_due': float(final_due),
                'mobile_number': student.mobile_number or '',
                'email': student.email or '',
                'student_id': student.id
            })()
            
            raw_data.append(row)
        
        return cls._organize_fee_report_by_class(raw_data)
    
    @classmethod
    def _organize_fee_report_by_class(cls, report_data):
        """Organize fee report data by class"""
        try:
            from core.utils import extract_class_number
            class_sort_key = extract_class_number
        except ImportError:
            class_sort_key = str
        
        # Group by class
        class_groups = defaultdict(list)
        for row in report_data:
            class_name = getattr(row, 'class_name', 'Unassigned')
            class_groups[class_name].append(row)
        
        # Sort classes and students within each class
        organized_data = []
        sorted_classes = sorted(class_groups.keys(), key=class_sort_key)
        
        for class_name in sorted_classes:
            students = sorted(class_groups[class_name], key=lambda x: getattr(x, 'name', ''))
            organized_data.extend(students)
        
        return organized_data
    
    @classmethod
    def _write_fees_report_csv(cls, writer, report_data):
        """Write fee report data to CSV organized by class"""
        current_class = None
        
        for row in report_data:
            class_name = getattr(row, 'class_name', 'Unassigned')
            
            # Add class header when class changes
            if current_class != class_name:
                if current_class is not None:  # Add empty row between classes
                    writer.writerow([])
                writer.writerow([f'=== {class_name} ===', '', '', '', '', '', '', '', '', '', '', '', ''])
                current_class = class_name
            
            writer.writerow([
                getattr(row, 'name', ''),
                getattr(row, 'admission_number', ''),
                getattr(row, 'class_name', ''),
                getattr(row, 'current_fees', 0),
                getattr(row, 'current_paid', 0),
                getattr(row, 'current_discount', 0),
                getattr(row, 'cf_due', 0),
                getattr(row, 'fine_paid', 0),
                getattr(row, 'fine_unpaid', 0),
                getattr(row, 'final_due', 0),
                'Paid' if getattr(row, 'final_due', 0) == 0 else 'Outstanding',
                getattr(row, 'mobile_number', ''),
                getattr(row, 'email', '')
            ])

# Legacy compatibility
ExportService = DataExportService