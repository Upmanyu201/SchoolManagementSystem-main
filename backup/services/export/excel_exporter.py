# backup/services/export/excel_exporter.py
"""Excel export functionality"""

import logging
import time
from datetime import datetime
from django.http import HttpResponse
from django.core.cache import cache
from django.core.exceptions import PermissionDenied, ValidationError

from .base import DataExportService
from .constants import ExportConstants
from .row_formatters import RowFormatter

export_logger = logging.getLogger('export.api')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

class ExcelExporter(DataExportService):
    """Excel export service"""
    
    @classmethod
    def export_to_excel(cls, module_name, user):
        """Export module data to Excel format with enhanced formatting"""
        start_time = time.time()
        export_logger.info(f"Excel export started for {module_name} by {user.username}")
        
        try:
            cls._validate_user_permissions(user, module_name)
            
            if not EXCEL_AVAILABLE:
                from .csv_exporter import CSVExporter
                return CSVExporter.export_to_csv(module_name, user)
            
            # Get fresh data for Excel export (avoid caching complex objects)
            data = cls.get_module_data(module_name)
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = cls._sanitize_filename(f"student_fee_report_{timestamp}.xlsx")
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Content-Transfer-Encoding'] = 'binary'
            response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response['Pragma'] = 'no-cache'
            response['Expires'] = '0'
            response['X-Content-Type-Options'] = 'nosniff'
            
            wb = Workbook()
            
            # Enhanced formatting for fees_report with class separation
            if module_name == 'fees_report':
                ws = wb.active
                ws.title = "Student Fee Report"
                
                # Add title row
                title_cell = ws.cell(row=1, column=1, value="STUDENT FEE REPORT")
                title_cell.font = Font(bold=True, size=16, color="FFFFFF")
                title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                ws.merge_cells('A1:I1')
                
                # Add headers with enhanced formatting
                headers = ExportConstants.CSV_HEADERS.get(module_name, ['Data'])
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=3, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                    
                    # Add borders to headers
                    thick_border = Border(
                        left=Side(style='thick'),
                        right=Side(style='thick'),
                        top=Side(style='thick'),
                        bottom=Side(style='thick')
                    )
                    cell.border = thick_border
                
                # Add data with class separation
                current_row = 4
                current_class = None
                
                for item in data:
                    try:
                        item_class = getattr(item, 'class_name', 'Unassigned')
                        
                        # Add class separator row
                        if current_class != item_class:
                            if current_row > 4:  # Add spacing between classes
                                current_row += 1
                            
                            # Class header row
                            class_cell = ws.cell(row=current_row, column=1, value=f"CLASS: {item_class}")
                            class_cell.font = Font(bold=True, size=12, color="FFFFFF")
                            class_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                            ws.merge_cells(f'A{current_row}:I{current_row}')
                            current_row += 1
                            current_class = item_class
                        
                        # Add student data row with formatting
                        row_data = cls._get_excel_row_safe(module_name, item)
                        for col_idx, value in enumerate(row_data, 1):
                            cell = ws.cell(row=current_row, column=col_idx, value=value)
                            
                            # Add borders
                            thin_border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                            cell.border = thin_border
                            
                            # Alternate row colors for better readability
                            if current_row % 2 == 0:
                                cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                            
                            # Right align numeric columns (amounts)
                            if col_idx >= 4 and col_idx <= 9:  # Amount columns (adjusted for removed context_type)
                                cell.alignment = Alignment(horizontal="right")
                            else:
                                cell.alignment = Alignment(horizontal="left")
                        
                        current_row += 1
                        
                    except Exception as e:
                        export_logger.warning(f"Error processing row in Excel export: {e}")
                        continue
            else:
                # Standard format for other modules
                ws = wb.active
                ws.title = f"{module_name.title()} Export"
                
                # Add headers with formatting
                headers = ExportConstants.CSV_HEADERS.get(module_name, ['Data'])
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # Pre-load fee report data for students
                if module_name == 'students' and data:
                    fee_report_data = cls._get_fee_report_data()
                    cls._fee_report_cache = {row.student_id: row.final_due for row in fee_report_data}
                
                # Add all data rows
                current_row = 2
                for item in data:
                    try:
                        row_data = cls._get_excel_row_safe(module_name, item)
                        
                        # Handle subjects module which returns multiple rows per item
                        if module_name == 'subjects' and isinstance(row_data, list) and len(row_data) > 0 and isinstance(row_data[0], list):
                            # Multiple rows returned
                            for row in row_data:
                                for col_idx, value in enumerate(row, 1):
                                    ws.cell(row=current_row, column=col_idx, value=value)
                                current_row += 1
                        else:
                            # Single row returned
                            for col_idx, value in enumerate(row_data, 1):
                                ws.cell(row=current_row, column=col_idx, value=value)
                            current_row += 1
                    except Exception as e:
                        export_logger.warning(f"Error processing row in Excel export: {e}")
                        continue
            
            # Clean up cache
            if hasattr(cls, '_fee_report_cache'):
                delattr(cls, '_fee_report_cache')
            
            # Auto-adjust column widths with better sizing
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                # Better width calculation for readability
                adjusted_width = min(max(max_length + 3, 10), 35)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save directly to response to avoid BytesIO corruption
            wb.save(response)
            
            total_time = time.time() - start_time
            export_logger.info(f"Excel export completed for {module_name} in {total_time:.2f}s")
            
            return response
            
        except (PermissionDenied, ValidationError) as e:
            export_logger.warning(f"Excel export denied for {module_name}: {e}")
            return HttpResponse(str(e), status=403)
        except Exception as e:
            export_logger.error(f"Excel export failed for {module_name}: {e}")
            return HttpResponse(f'Excel export failed: {str(e)}', status=500)
    
    @classmethod
    def _get_excel_row_safe(cls, module_name, item):
        """Get Excel row data with safe attribute access"""
        try:
            if module_name == 'students':
                # Get final due from centralized fee report calculation
                final_due = 0
                try:
                    student_id = getattr(item, 'id', None)
                    if student_id and hasattr(cls, '_fee_report_cache'):
                        final_due = cls._fee_report_cache.get(student_id, 0)
                    elif student_id:
                        fee_report_data = cls._get_fee_report_data()
                        for fee_row in fee_report_data:
                            if getattr(fee_row, 'student_id', None) == student_id:
                                final_due = getattr(fee_row, 'final_due', 0)
                                break
                except Exception:
                    final_due = 0
                
                return RowFormatter.format_student_row(item, final_due)
            elif module_name == 'teachers':
                return RowFormatter.format_teacher_row(item)
            elif module_name == 'subjects':
                # Handle multiple rows returned by format_subject_row
                rows = RowFormatter.format_subject_row(item)
                return rows  # Return all rows for this class section
            elif module_name == 'transport':
                return RowFormatter.format_transport_row(item)
            elif module_name == 'fees':
                return RowFormatter.format_fees_row(item)
            elif module_name == 'student_fees':
                return RowFormatter.format_student_fees_row(item)

            elif module_name == 'attendance':
                return RowFormatter.format_attendance_row(item)


            elif module_name == 'users':
                return RowFormatter.format_users_row(item)
            elif module_name == 'fees_report':
                return RowFormatter.format_fees_report_row(item)
            else:
                return [str(item)]
                
        except Exception as e:
            export_logger.error(f"Error processing Excel row for {module_name}: {e}")
            return ['Error processing data']