
from django.core.cache import cache
from django.http import HttpResponse
from django.views.decorators.cache import cache_page
from django.views.decorators.vary import vary_on_headers
import threading
import time

# Thread-local storage for request tracking
_local = threading.local()

def optimize_export_performance(view_func):
    """Decorator to optimize export performance"""
    def wrapper(request, *args, **kwargs):
        # Quick response for concurrent requests
        start_time = time.time()
        
        # Cache key based on request parameters
        cache_key = f"export_{request.path}_{request.GET.urlencode()}"
        
        # Check if similar request is being processed
        processing_key = f"processing_{cache_key}"
        if cache.get(processing_key):
            # Return cached result if available
            cached_result = cache.get(cache_key)
            if cached_result:
                return cached_result
        
        # Mark as processing
        cache.set(processing_key, True, 30)  # 30 seconds timeout
        
        try:
            response = view_func(request, *args, **kwargs)
            
            # Cache successful responses for 5 minutes
            if response.status_code == 200:
                cache.set(cache_key, response, 300)
            
            return response
        finally:
            # Clear processing flag
            cache.delete(processing_key)
            
            # Log performance
            duration = time.time() - start_time
            if duration > 5:  # Log slow requests
                print(f"Slow export: {request.path} took {duration:.2f}s")
    
    return wrapper

def create_lightweight_response(data, format_type, filename):
    """Create optimized response for large datasets"""
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        
        import csv
        writer = csv.writer(response)
        
        # Write headers
        if data:
            writer.writerow(data[0].keys())
            
            # Write data in chunks to avoid memory issues
            for record in data:
                writer.writerow(record.values())
        
        return response
    
    elif format_type == 'excel':
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        if data:
            # Headers
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Data
            for row, record in enumerate(data, 2):
                for col, value in enumerate(record.values(), 1):
                    ws.cell(row=row, column=col, value=value)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        
        wb.save(response)
        return response
    
    return HttpResponse("Unsupported format", status=400)

# Connection pooling for database queries
class DatabaseOptimizer:
    @staticmethod
    def optimize_queryset(queryset):
        """Optimize queryset for export"""
        return queryset.select_related().prefetch_related()
    
    @staticmethod
    def batch_process(queryset, batch_size=1000):
        """Process queryset in batches"""
        count = queryset.count()
        for i in range(0, count, batch_size):
            yield queryset[i:i + batch_size]
